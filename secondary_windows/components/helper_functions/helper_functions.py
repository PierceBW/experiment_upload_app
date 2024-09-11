#Can be deleted

# %%
import pandas as pd
import re
from pint import UnitRegistry
import os
import json
ureg = UnitRegistry()
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import boto3
import os
from dotenv import load_dotenv
import numpy as np

# %% [markdown]
# # Helper Functions
#
# 

# %% [markdown]
# ## Plate Map Overvitreatment2
# Lets look at what is included in a plate map

# %%
# Log in and get S3
load_dotenv()

client = boto3.client(
    's3',
    aws_access_key_id=os.getenv('aws_access_key_id'),
    aws_secret_access_key=os.getenv('aws_secret_access_key'),
    aws_session_token=os.getenv('aws_session_token')
)

# %%
# Run this chunk to get all experiments
def list_metadata_files(bucket_name):
    s3 = client
    
    # Use list_objects_v2 with Delimiter to list folders (common prefixes)
    response = s3.list_objects_v2(Bucket=bucket_name, Delimiter='/')

    if 'CommonPrefixes' in response:
        experiments = []
        for prefix in response['CommonPrefixes']:
            # Get the prefix (folder name)
            folder_name = prefix['Prefix']
            # List objects in the folder (one level deep)
            objects_response = s3.list_objects_v2(Bucket=bucket_name, Prefix=folder_name, Delimiter='/')
            temp = [obj['Prefix'] for obj in objects_response['CommonPrefixes']]
            experiments += temp
    else:
        print("No folders found in the bucket {}.".format(bucket_name))

    return experiments

# Specify the bucket name
bucket_name = os.getenv('bucket')

# Call the function to list metadata files one folder deep
experiments = list_metadata_files(bucket_name)

# Run this to get necessary functions
def load_json_from_s3(bucket_name, key):
    s3_client = client
    try:
        response = s3_client.get_object(Bucket=bucket_name, Key=key)
        json_data = response['Body'].read().decode('utf-8')
        json_object = json.loads(json_data)
        return json_object
    except Exception as e:
        print(f"Error loading JSON file from S3: {e}")
        return None
    

# %%
# get the plate map
plate_1 = edu_metadata['plate_data']['1']['plate_map']
print(plate_1.keys())

plate_1 = mft_metadata['plate_data']['1']['plate_map']
print(plate_1.keys())
example_df = pd.DataFrame(plate_1)

# %% [markdown]
# ## Implement plate map creation

# %%

def get_serum_val_edu(df, type_value):
    """Gets corresponding serum value to treatment type"""
    if "control" in type_value:
        serum_value = df.loc[df["Type"] == type_value, "Serum"]
        return serum_value.iloc[0] * 100.0
    elif pd.isna(type_value) or not type_value:
         return pd.NA
    else:
        serum_value = df.loc[df["Type"] == "treatment", "Serum"]
        return serum_value.iloc[0] * 100.0
    
def get_treatment_val_mft(df, type_value):
    """Gets corresponding treatment1 and treatment2 values for treatment type"""
    treatment1 = df.loc[df["standard"] == "treatment1", "concentration"]
    treatment1 = treatment1.iloc[0]
    treatment2 = df.loc[df["standard"] == "treatment2", "concentration"]
    treatment2 = treatment2.iloc[0]

    treatment1 = float(re.findall(r'[.\d]+', treatment1)[0])
    treatment2 = float(re.findall(r'[.\d]+', treatment2)[0])

    if "positive control" == type_value:
        return [treatment1, treatment2]
    elif pd.isna(type_value) or not type_value:
        return [pd.NA, pd.NA]
    elif "untreated control" == type_value:
        return [0, 0]
    else:
        return [treatment1, 0]


def mft_standards_columns(multiplate_map_df, standards_df, experiment_units):
    """
    Add the treatment columns to the multiplate map
    """
    mft_df = multiplate_map_df["treatment"].apply(lambda type: get_treatment_val_mft(standards_df, type))
    mft_df = pd.DataFrame(mft_df.tolist(), columns=["treatment1_concentration (" + experiment_units + ")", "treatment2_concentration (" + experiment_units + ")"])
    final_df = pd.concat([multiplate_map_df, mft_df], axis=1)
    final_df = final_df.iloc[:, [0,1,2,3,4,7,6,5]]
    return final_df

def edu_standards_columns(multiplate_map_df, standards_df, experiment_units):
    """
    Takes treatment data from dataframe and adds corresponding serum percentage to dataframe
    """
    edu_df = multiplate_map_df["treatment"].apply(lambda type: get_serum_val_edu(standards_df, type))
    edu_df = pd.DataFrame(edu_df.tolist(), columns=["serum (percentage)"])
    final_df = pd.concat([multiplate_map_df, edu_df], axis=1)
    return final_df

def convert_combine(dict1, df1):
    """Converts df to dictionary to combine"""
    dictp = df1.to_dict('list')
    for key in dict1:
            if key in dictp:
                dict1[key].extend(dictp[key])
    for key in dictp:
            if key not in dict1:
                dict1[key] = dictp[key]

def convert_to_multiplate_map(plate_maps_path, experiment_type, experiment_units):
    """
    Converts a plate map to from wide to long format and adds rows to specify treatment concentrations

    inputs: path to plate map to convert and the type of experiment

    output: dataframe in long format with treatment data
    """
    dict = {}
    plate_map = pd.read_excel(plate_maps_path, sheet_name=None)
    for key, value in plate_map.items():
        if key == "standards":
            pass
        else:
            #turn from wide to long format
            df = value.melt(id_vars = "row", var_name = "column")
            ntreatment2_df = df['value'].apply(lambda val: pd.Series(str(val).split("\n")))
            df = pd.concat([df[['row', 'column']], ntreatment2_df], axis=1)
            df = df.rename({0: "treatment", 1: "treatment_concentration"}, axis = "columns")
            
            #add row for well and plate
            df["well"] = df["row"] + df["column"].astype(str)
            df['plate'] = int(key.split("plate")[1])
            df = df[["row", "column", "treatment", "plate", "well","treatment_concentration"]]
            
            #Turn df into dict to add all plate maps together
            convert_combine(dict, df)
    
    #convert dict back to dataframe
    df_final = pd.DataFrame.from_dict(dict)

    #Dict for experiemnt type and their coresponding function to call
    standards = {"mft": mft_standards_columns, "edu": edu_standards_columns}
    
    if experiment_type not in standards:
            ExperimentTypeNotImplementedError(experiment_type)
    else:
         pass
    #Add rows specific to experiment type
    df_final = standards[experiment_type](df_final, plate_map["standards"], experiment_units)
    #fix and remove possible units in treatment_concentration
    df_final["treatment_concentration"] = df_final.apply(lambda x: 0 if "control" in x["treatment"] else x['treatment_concentration'], axis=1)
    df_final["treatment_concentration"] = df_final["treatment_concentration"].apply(lambda val: (re.findall(r'[.\d]+', val)[0]) if isinstance(val, str) else val)
    df_final = df_final.rename({"treatment_concentration": "treatment_concentration (" + experiment_units + ")"}, axis = "columns")
    
    return df_final

# %%
#testdf = convert_to_multiplate_map('helper_functions_example_files/mft_example_plate_map.xlsx', 'mft', "ng/ml")
#testdf

# %% [markdown]
# ## Name correction and molecular weight mapper
# These functions will take our current list of peptides and create a dataframe with primary names, secondary names, and molecular weight

# %%

# %%
def correct_name(row:pd.DataFrame):
    """
    Helper function for update_multiplate_map. This will find the primary name for a row in plate map
    """
    if 'control' in row['treatment']:
        row['treatment_updated'] = row['treatment']
        row['MW (daltons)'] = pd.NA
    else:
        row['treatment_updated'] = row['primary name_x']
        row['MW (daltons)'] = row['MW (daltons)_x']
        if pd.isna(row['treatment_updated']):
            row['treatment_updated'] = row['primary name_y']
            row['MW (daltons)'] = row['MW (daltons)_y']
        if pd.isna(row['treatment_updated']):
            treatment = row['treatment']
            print(f'Missing {treatment}')
    #row = row.drop(columns=['primary name_x','secondary names_x','MW (daltons)_x','primary name_y','secondary names_y','MW (daltons)_y'])
    return row

def update_multiplate_map(multiplate_map_df, molecular_weight_df_path = "secondary_windows/components/helper_functions/molecular_weight_df_empty.pkl"):
    """
    Correct peptide names in plate map

    Args:
        multiplate_map_df (pandas.DataFrame): plate map with treatment column
        molecular_weight_df (pandas.DataFrame): molecular weight df with primary and secondar names

    Returns:
        pandas.Dataframe: dataframe with updated peptide names
    """
    molecular_weight_df = pd.read_pickle(molecular_weight_df_path)

    multiplate_map_df = multiplate_map_df.merge(molecular_weight_df,left_on='treatment',right_on='primary name',how='left')
    multiplate_map_df = multiplate_map_df.merge(molecular_weight_df.explode('secondary names'),left_on='treatment',right_on='secondary names',how='left')
    multiplate_map_df = multiplate_map_df.apply(correct_name,axis=1).drop(columns=['primary name_x','secondary names_x','MW (daltons)_x','primary name_y','secondary names_y','MW (daltons)_y', 'treatment'])
    multiplate_map_df = multiplate_map_df.rename({"treatment_updated": "treatment"}, axis = "columns")
    # Insert the 'treatment' column at the third position (index 2)
    cols = multiplate_map_df.columns.tolist()
    cols.remove('treatment')
    cols.insert(2, 'treatment')
    multiplate_map_df = multiplate_map_df[cols]
    return multiplate_map_df

def convert_concentration(concentration,units,molecular_weight,target_units):
    """
    Given a concentration in molarity or g/l convert to target units

    Args:
        concentration (float): numerical value of concentration
        units (str): units of concentrations
        molecular_weight (float): numerical value of molecular weight in grams/mole
        target_units (str): target units of conversion
    
    Returns:
        (float): concentration in target units
    """
    ureg.define('molarity = mol/L = M')
    units = ureg(units)
    concentration = concentration * units
    molecular_weight = molecular_weight * ureg.gram / ureg.mole

    if units.to_base_units().units == ureg('mole/meter**3').units:
        concentration = concentration * molecular_weight
    else:
        concentration = concentration / molecular_weight
    
    return round(concentration.to(target_units).magnitude,2)

convert_concentration(20,'ng/ml',24000,'nM')
def update_concentrations(multiplate_map_df, experiment_type, treatment1_mw=24000.,treatment2_mw=399.42):
    """
    Function to add missing concentration type

    Args:
        multiplate_map_df (pandas.DataFrame): Plate map in row form holding all plates from a given experiment
        treatment1_mw (float): molecular weight of treatment1 in g/mol
        etreatment2_mw (float): molecular weight of treatment2 in g/mol

    Returns:
        pandas.DataFrame: return multiplate map with missing concentration types
    """
    # I want you to implement this function so missing concetration types are dynamically added.
    # If a concentration column is ng/ml create one for nM or vice versa
    unit_list = ["nM", "ng/ml"]
    #List of columns
    col = multiplate_map_df.columns.tolist()
    #Find the units
    for string in col:
        if "treatment_concentration" in string:
            units = re.findall(r'\((.*?)\)', string)[0]
    #Get opposite units
    unit_list.remove(units)
    oppUnits = unit_list[0]
    
    if experiment_type == "mft":
        # dd other treatment1 column
        multiplate_map_df["treatment2_concentration (" + oppUnits + ")"] = multiplate_map_df.apply(lambda x: convert_concentration(float(x["treatment2_concentration (" + units + ")"]), units, treatment2_mw, oppUnits), axis=1)
        #add other treatment2 column
        multiplate_map_df["treatment1_concentration (" + oppUnits + ")"] = multiplate_map_df.apply(lambda x: convert_concentration(float(x["treatment1_concentration (" + units + ")"]), units, treatment1_mw, oppUnits), axis=1)
    else:
        pass
    #add other treatment concentration column
    multiplate_map_df["treatment_concentration (" + oppUnits + ")"] = multiplate_map_df.apply(lambda x: 0 if "control" in x["treatment"] else convert_concentration(float(x["treatment_concentration (" + units + ")"]), units, x["MW (daltons)"], oppUnits), axis=1)
    
    
    return multiplate_map_df

# %%
#testdf1 = convert_to_multiplate_map('helper_functions_example_files/mft_example_plate_map.xlsx', 'mft', "nM")
#testdf1 = update_multiplate_map(testdf1, molecular_weight_df)

#testdf2 = convert_to_multiplate_map('helper_functions_example_files/edu_example_plate_map.xlsx', 'edu', "ng/ml")
#testdf2 = update_multiplate_map(testdf2, molecular_weight_df)

#convert_concentration(concentration,units,molecular_weight,target_units)
#update_concentrations(multiplate_map_df, experiment_type, treatment1_mw=24000.,treatment2_mw=399.42)
#update_concentrations(testdf1, "mft")
#testdf1

# %% [markdown]
# ## Create human readable plate maps from plate map

# %%
def combine_treatment_information(row,control_dict):
    """
    This creates a string with all the necessary treatment information

    Args:
        row (pandas.DataFrame): This is a row in the plate map
        control_dict (dict): If the row is a control convert the controls name using the control_dict

    Returns:
        str: string with all the treatment information (treatment name and concentration in both ng/ml and nM)
    """
    treatment = row['treatment']
    if row['treatment'] in control_dict.keys():
        return control_dict[row['treatment']]
    conc_ngml = row['treatment_concentration (ng/ml)']
    conc_nM = row['treatment_concentration (nM)']
    value = f'{treatment}\n{conc_ngml}ng/mL\n{conc_nM}nM'
    return value

def format_excel(spreadsheet_name):
    """
    This formats an excel spreadsheet so it's easy to ready and understand

    Args:
        spreadsheet_name (str): path to spreadsheet to be formatted
    """
    wb = load_workbook(spreadsheet_name)
    ws = wb.active

    # Apply text wrapping and adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
                cell.alignment = Alignment(wrap_text=True)  # Apply text wrapping
            except:
                pass
        adjusted_width = (max_length) / 2  # Adjust the multiplier as needed
        if column == 'A':
            ws.column_dimensions[column].width = max_length +2
        else:
            ws.column_dimensions[column].width = adjusted_width
    # Save the modified Excel file
    wb.save(spreadsheet_name)
    wb.close()

def mft_control_quanitity(plate_map_df,plate_map_df_old):
    """
    Adds description column with control information for the mft assay

       Args:
        plate_map_df (pandas.DataFrame): Matrix version of plate map in a pandas dataframe
        plate_map_df_old (pandas.DataFrame): row version of plate map in pandas dataframe

    Returns:
        pandas.DataFrame: returns a matrix version of the plate map in a pandas dataframe with an additional column for control information
    """
    plate_map_df['Controls'] = pd.NA
    control_concentrations = {'treatment1':[plate_map_df_old['treatment1_concentration (ng/ml)'].max(),plate_map_df_old['treatment1_concentration (nM)'].max()],'treatment2':[plate_map_df_old['treatment2_concentration (ng/ml)'].max(),plate_map_df_old['treatment2_concentration (nM)'].max()]}
    for idx, (key, value) in enumerate(control_concentrations.items()):
        plate_map_df.loc[idx, 'Controls'] = f'{key}\n{value[0]}ng/ml\n{value[1]}nM'
    return plate_map_df

def edu_serum_quantity(plate_map_df,plate_map_df_old):
    """
    Adds description column with serum information for the mft assay

    Args:
        plate_map_df (pandas.DataFrame): Matrix version of plate map in a pandas dataframe
        plate_map_df_old (pandas.DataFrame): row version of plate map in pandas dataframe

    Returns:
        pandas.DataFrame: returns a matrix version of the plate map in a pandas dataframe with an additional column for serum information
    """
    plate_map_df['Serum'] = pd.NA
    serum_percentages = plate_map_df_old['serum (percentage)'].unique()
    serum_percentages = {'positive control':serum_percentages.max(),'negative control':np.median(serum_percentages),'treatment':serum_percentages.min()}
    for idx, (key, value) in enumerate(serum_percentages.items()):
        plate_map_df.loc[idx, 'Serum'] = f'{key}\n{value}%'
    return plate_map_df

class ExperimentTypeNotImplementedError(Exception):
    def __init__(self,experiment_type,message='This experiment type is not impletented yet.'):
        self.experiment_type = experiment_type
        self.message = f'{message}. Experiment Type: {self.experiment_type}'
        super().__init__(self.message)

control_dict = {'positive control':'positive control\n(treatment2 & treatment1)','negative control':'negative control\n(treatment1)'}
def create_human_readable_plate_map(plate_map_df,plate_full_name,experiment_type,control_dict):
    """
    Creates a human readable plate map from a plate map dataframe in row form

    Args:
        plate_map_df (pandas.DataFrame): plate map in row form
        plate_full_name (str): name to save spreadsheet under
        experiment_type (str): name of experiment type for experiment specific processing
        control_dict (dict): dictionary for renaming controls

    Return:
        str: spreadsheet name with file suffic
    """
    experiment_specialized_functions = {'mft':mft_control_quanitity,'edu':edu_serum_quantity}
    print('formatting treatment info')
    plate_map_df_old = plate_map_df.copy()
    plate_map_df['value'] = plate_map_df.apply(combine_treatment_information,args=(control_dict,),axis=1)
    plate_map_df = plate_map_df[['row','column','value']]
    print('transforming df for readability')
    plate_map_df = plate_map_df.pivot(index='row',columns='column',values='value').reset_index()
    if experiment_type in experiment_specialized_functions:
        plate_map_df = experiment_specialized_functions[experiment_type](plate_map_df,plate_map_df_old)
    else:
        ExperimentTypeNotImplementedError(experiment_type)
    spreadsheet_name = f'{plate_full_name}.xlsx'
    plate_map_df.to_excel(spreadsheet_name,index=False)
    format_excel(spreadsheet_name)
    return spreadsheet_name

def cleanup(spreadsheet_names):
    """
    Function to remove spreadsheets from local directory after uploading to S3

    Args:
        spreadsheet_names(list[str]): list of spreadsheet path names
    """
    for file in spreadsheet_names:
            os.remove(file)

# %%
def save_experiment_data(multiplate_map_df,experiment_data,s3_client,bucket_name,control_dict=control_dict):
    experiment_path = experiment_data['experiment_folder_path']
    for plate, data in experiment_data['plate_data'].items():
        plate_map_df = multiplate_map_df.loc[multiplate_map_df['plate'] == plate]
        experiment_data['plate_data'][plate]['plate_map'] = plate_map_df.to_dict()
        spreadsheet_name = create_human_readable_plate_map(plate_map_df,data['plate_full_name'],experiment_data['experiment_type'],control_dict)
        s3_key = f"{experiment_path}plate_maps/{spreadsheet_name}"
        s3_client.upload_file(spreadsheet_name, bucket_name, s3_key)
    
    json_content = json.dumps(experiment_data)
    s3_client.put_object(Bucket=bucket_name, Key=f'{experiment_path}metadata.json', Body=json_content)

    return experiment_data


