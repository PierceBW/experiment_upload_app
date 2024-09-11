#Create human readbale plate maps functions
import os
import pandas as pd
import numpy as np
import json
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def combine_treatment_information(row,control_dict):
    """
    This creates a string with all the necessary treatment information

    Args:
        row (pandas.DataFrame): This is a row in the plate map
        control_dict (dict): If the row is a control convert the controls name using the control_dict

    Returns:
        str: string with all the treatment information (treatment name and concentration in both ng/ml and nM)
    """
    treatment = row['treatment']
    if row['treatment'] in control_dict.keys():
        return control_dict[row['treatment']]
    conc_ngml = row['treatment_concentration (ng/ml)']
    conc_nM = row['treatment_concentration (nM)']
    value = f'{treatment}\n{conc_ngml}ng/mL\n{conc_nM}nM'
    return value

def format_excel(spreadsheet_name):
    """
    This formats an excel spreadsheet so it's easy to ready and understand

    Args:
        spreadsheet_name (str): path to spreadsheet to be formatted
    """
    wb = load_workbook(spreadsheet_name)
    ws = wb.active

    # Apply text wrapping and adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
                cell.alignment = Alignment(wrap_text=True)  # Apply text wrapping
            except:
                pass
        adjusted_width = (max_length) / 2  # Adjust the multiplier as needed
        if column == 'A':
            ws.column_dimensions[column].width = max_length +2
        else:
            ws.column_dimensions[column].width = adjusted_width
    # Save the modified Excel file
    wb.save(spreadsheet_name)
    wb.close()

def mft_control_quanitity(plate_map_df,plate_map_df_old):
    """
    Adds description column with control information for the mft assay

       Args:
        plate_map_df (pandas.DataFrame): Matrix version of plate map in a pandas dataframe
        plate_map_df_old (pandas.DataFrame): row version of plate map in pandas dataframe

    Returns:
        pandas.DataFrame: returns a matrix version of the plate map in a pandas dataframe with an additional column for control information
    """
    plate_map_df['Controls'] = pd.NA
    control_concentrations = {'treatment1':[plate_map_df_old['treatment1_concentration (ng/ml)'].max(),plate_map_df_old['treatment1_concentration (nM)'].max()],'treatment2':[plate_map_df_old['treatment2_concentration (ng/ml)'].max(),plate_map_df_old['treatment2_concentration (nM)'].max()]}
    for idx, (key, value) in enumerate(control_concentrations.items()):
        plate_map_df.loc[idx, 'Controls'] = f'{key}\n{value[0]}ng/ml\n{value[1]}nM'
    return plate_map_df

def edu_serum_quantity(plate_map_df,plate_map_df_old):
    """
    Adds description column with serum information for the mft assay

    Args:
        plate_map_df (pandas.DataFrame): Matrix version of plate map in a pandas dataframe
        plate_map_df_old (pandas.DataFrame): row version of plate map in pandas dataframe

    Returns:
        pandas.DataFrame: returns a matrix version of the plate map in a pandas dataframe with an additional column for serum information
    """
    plate_map_df['Serum'] = pd.NA
    serum_percentages = plate_map_df_old['serum (percentage)'].unique()
    serum_percentages = {'positive control':serum_percentages.max(),'negative control':np.median(serum_percentages),'treatment':serum_percentages.min()}
    for idx, (key, value) in enumerate(serum_percentages.items()):
        plate_map_df.loc[idx, 'Serum'] = f'{key}\n{value}%'
    return plate_map_df

class ExperimentTypeNotImplementedError(Exception):
    def __init__(self,experiment_type,message='This experiment type is not impletented yet.'):
        self.experiment_type = experiment_type
        self.message = f'{message}. Experiment Type: {self.experiment_type}'
        super().__init__(self.message)

control_dict = {'positive control':'positive control\n(treatment2 & treatment1)','negative control':'negative control\n(treatment1)'}
def create_human_readable_plate_map(plate_map_df,plate_full_name,experiment_type,control_dict):
    """
    Creates a human readable plate map from a plate map dataframe in row form

    Args:
        plate_map_df (pandas.DataFrame): plate map in row form
        plate_full_name (str): name to save spreadsheet under
        experiment_type (str): name of experiment type for experiment specific processing
        control_dict (dict): dictionary for renaming controls

    Return:
        str: spreadsheet name with file suffic
    """
    experiment_specialized_functions = {'mft':mft_control_quanitity,'edu':edu_serum_quantity}
    print('formatting treatment info')
    plate_map_df_old = plate_map_df.copy()
    plate_map_df['value'] = plate_map_df.apply(combine_treatment_information,args=(control_dict,),axis=1)
    plate_map_df = plate_map_df[['row','column','value']]
    print('transforming df for readability')
    plate_map_df = plate_map_df.pivot(index='row',columns='column',values='value').reset_index()
    if experiment_type in experiment_specialized_functions:
        plate_map_df = experiment_specialized_functions[experiment_type](plate_map_df,plate_map_df_old)
    else:
        ExperimentTypeNotImplementedError(experiment_type)
    spreadsheet_name = f'{plate_full_name}.xlsx'
    plate_map_df.to_excel(spreadsheet_name,index=False)
    format_excel(spreadsheet_name)
    return spreadsheet_name

def cleanup(spreadsheet_names):
    """
    Function to remove spreadsheets from local directory after uploading to S3

    Args:
        spreadsheet_names(list[str]): list of spreadsheet path names
    """
    for file in spreadsheet_names:
            os.remove(file)

#
def save_experiment_data(multiplate_map_df,experiment_data,s3_client,bucket_name,control_dict=control_dict):
    experiment_path = experiment_data['experiment_folder_path']
    for plate, data in experiment_data['plate_data'].items():
        plate_map_df = multiplate_map_df.loc[multiplate_map_df['plate'] == plate]
        experiment_data['plate_data'][plate]['plate_map'] = plate_map_df.to_dict()
        spreadsheet_name = create_human_readable_plate_map(plate_map_df,data['plate_full_name'],experiment_data['experiment_type'],control_dict)
        s3_key = f"{experiment_path}plate_maps/{spreadsheet_name}"
        s3_client.upload_file(spreadsheet_name, bucket_name, s3_key)
    
    json_content = json.dumps(experiment_data)
    s3_client.put_object(Bucket=bucket_name, Key=f'{experiment_path}metadata.json', Body=json_content)

    return experiment_data
